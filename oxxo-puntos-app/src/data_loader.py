"""
Carga y normalización de las bases de datos de la aplicación.

- ``Book.xlsx`` (hoja ``JUN``): tiendas vigentes con coordenadas X/Y.
- ``Operaciones_ult_semana.xlsm`` (hoja ``Visitas_Operaciones``): puntos
  evaluados. Se usan las coordenadas de las columnas Y (latitud) y X (longitud)
  si existen; si no, se extraen del enlace de Google Maps usando la cascada
  de maps_utils.py.
"""
from __future__ import annotations

import os
from typing import Iterable

import pandas as pd
import streamlit as st

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
BOOK_PATH = os.path.join(BASE_DIR, "data", "Book.xlsx")
VISITAS_PATH = os.path.join(BASE_DIR, "data", "Operaciones_ult_semana.xlsm")

ESTADOS_VIGENTES = ["ABIERTA", "OBRA", "FIRMADA"]

TIENDAS_COLS = [
    "NAME", "ESTADO", "PLAZA 2026", "DEPARTAMENTO", "MUNICIPIO",
    "UPZ/COMUNA", "ESTRATO", "TIPO DE LOCAL", "AREA", "X", "Y",
    "FECHA APE", "ARRENDADOR",
]

VISITAS_RENAME = {
    "Nombre del Punto ": "Nombre del Punto",
    " TICKET U6M": "TICKET U6M",
    " VENTAS OUM": "VENTAS OUM",
    " CONTRIBUCION UM": "CONTRIBUCION UM",
    " CONTRIBUCION U6M": "CONTRIBUCION U6M",
    " RENTA UM": "RENTA UM",
}

# Nombre de columna estándar para el filtro de fecha de la interfaz.
DATE_COLUMN_STD = "Fecha"
DATE_COLUMN_HINTS = ["fecha de visita", "fecha visita", "fecha de la visita", "fecha"]


def _file_signature(path: str):
    """mtime + size para invalidar el caché cuando se reemplaza un Excel."""
    stat = os.stat(path)
    return (path, stat.st_mtime, stat.st_size)


def _find_date_column(df: pd.DataFrame):
    """Encuentra una columna de fecha en ``Visitas_Operaciones``."""
    cols_lower = {c.lower().strip(): c for c in df.columns if isinstance(c, str)}
    for hint in DATE_COLUMN_HINTS:
        if hint in cols_lower:
            return cols_lower[hint]
    for lower, original in cols_lower.items():
        if "fecha" in lower:
            return original
    return None


def _normalizar_encabezado(value: object) -> str:
    """Normaliza un encabezado para comparar variantes de escritura."""
    return "".join(str(value).casefold().strip().replace("_", " ").split())


def _add_visit_coordinates(df: pd.DataFrame) -> tuple[str | None, str | None]:
    """
    Agrega las columnas normalizadas ``lat`` y ``lon``.

    Prioridad de fuentes:
    1. Columnas Y (latitud) y X (longitud) explícitas en el Excel
    2. Enlace de Google Maps (parseo + resolución de links cortos)
    3. Enlace de Bing Maps
    4. Geocodificación por dirección con Nominatim
    """
    df["lat"] = pd.Series(float("nan"), index=df.index, dtype="float64")
    df["lon"] = pd.Series(float("nan"), index=df.index, dtype="float64")

    maps_col = None
    address_col = None

    # --- FUENTE 1: Columnas Y (lat) y X (lon) explícitas ---
    tiene_y = "Y" in df.columns
    tiene_x = "X" in df.columns
    usar_xy = tiene_y and tiene_x

    if usar_xy:
        # Convertir Y y X a numérico
        df["lat"] = pd.to_numeric(df["Y"], errors="coerce")
        df["lon"] = pd.to_numeric(df["X"], errors="coerce")
        # Validar rango
        valid = df["lat"].between(-90, 90) & df["lon"].between(-180, 180)
        df.loc[~valid, ["lat", "lon"]] = float("nan")

    # --- FUENTE 2: Enlace de Google Maps ---
    # Buscar la columna de Maps
    for col in df.columns:
        if isinstance(col, str):
            col_lower = col.lower().strip()
            if "enlace" in col_lower and "maps" in col_lower:
                maps_col = col
                break
    if maps_col is None:
        for col in df.columns:
            if isinstance(col, str) and "maps" in col.lower():
                maps_col = col
                break

    # Buscar columna de dirección para respaldo
    for col in df.columns:
        if isinstance(col, str):
            col_lower = col.lower().strip()
            if "dirección" in col_lower or "direccion" in col_lower or "address" in col_lower:
                address_col = col
                break

    # Si la fuente 1 no cubrió todos, completar con Maps
    sin_coordenadas = df["lat"].isna()
    if maps_col and sin_coordenadas.any():
        # Primero intentamos leer hipervínculos de Excel directamente
        from openpyxl import load_workbook
        wb = load_workbook(VISITAS_PATH, read_only=True)
        ws = wb['Visitas_Operaciones']
        
        maps_col_idx = None
        id_col_idx = None
        for idx, col in enumerate(ws[1]):
            if isinstance(col.value, str) and ("maps" in col.value.lower() or "ubicación" in col.value.lower()):
                maps_col_idx = idx
            if isinstance(col.value, str) and "id" in col.value.lower():
                id_col_idx = idx
                
        if maps_col_idx is not None:
            for row in ws.iter_rows(min_row=2, values_only=False):
                cell = row[maps_col_idx]
                if cell.hyperlink and cell.hyperlink.target:
                    import re
                    target = cell.hyperlink.target
                    lat, lon = None, None
                    match = re.search(r'@(-?\d+\.\d+),(-?\d+\.\d+)', target)
                    if match:
                        lat, lon = float(match.group(1)), float(match.group(2))
                    match = re.search(r'/@(-?\d+\.\d+),(-?\d+\.\d+)', target)
                    if match:
                        lat, lon = float(match.group(1)), float(match.group(2))
                    match = re.search(r'!1d(-?\d+\.\d+)!2d(-?\d+\.\d+)', target)
                    if match:
                        lat, lon = float(match.group(1)), float(match.group(2))
                    match = re.search(r'!2d(-?\d+\.\d+)!3d(-?\d+\.\d+)', target)
                    if match:
                        lat, lon = float(match.group(1)), float(match.group(2))
                        
                    if lat is not None and lon is not None:
                        if id_col_idx is not None:
                            id_val = row[id_col_idx].value
                            if id_val is not None:
                                matching_idx = df[df['ID'] == str(id_val)].index
                                if len(matching_idx) > 0:
                                    df.at[matching_idx[0], "lat"] = lat
                                    df.at[matching_idx[0], "lon"] = lon
        wb.close()
        
        # Si aún hay sin coordenadas, usamos el método tradicional
        sin_coordenadas = df["lat"].isna()
        if sin_coordenadas.any():
            from src.maps_utils import get_coordinates, geocode_address

            for idx, row in df[sin_coordenadas].iterrows():
                maps_link = row.get(maps_col, "")
                address = row.get(address_col, "") if address_col else ""
                if not isinstance(maps_link, str) or not maps_link.strip():
                    if isinstance(address, str) and address.strip():
                        coords = geocode_address(address)
                        if coords:
                            df.at[idx, "lat"] = coords[0]
                            df.at[idx, "lon"] = coords[1]
                    continue
                lat, lon, _ = get_coordinates(
                    maps_link, address if isinstance(address, str) else ""
                )
                if lat is not None and lon is not None:
                    df.at[idx, "lat"] = lat
                    df.at[idx, "lon"] = lon

    # Validación final
    valid = df["lat"].between(-90, 90) & df["lon"].between(-180, 180)
    df.loc[~valid, ["lat", "lon"]] = float("nan")

    return maps_col, address_col


@st.cache_data(show_spinner="Cargando tiendas vigentes...")
def load_tiendas(_sig=None) -> pd.DataFrame:
    _file_signature(BOOK_PATH)
    df = pd.read_excel(BOOK_PATH, sheet_name="JUN")
    df.columns = [str(c).strip() for c in df.columns]
    keep = [c for c in TIENDAS_COLS if c in df.columns]
    df = df[keep].copy()

    df["ESTADO"] = df["ESTADO"].astype(str).str.strip().str.upper()
    df = df[df["ESTADO"].isin(ESTADOS_VIGENTES)].copy()

    df["NAME"] = df["NAME"].astype(str).str.strip()
    df = df[df["NAME"].ne("") & df["NAME"].ne("0")]

    # X = longitud, Y = latitud en la base de tiendas.
    df["lat"] = pd.to_numeric(df["Y"], errors="coerce")
    df["lon"] = pd.to_numeric(df["X"], errors="coerce")
    return df.reset_index(drop=True)


@st.cache_data(show_spinner="Cargando puntos evaluados (Operaciones)...")
def load_visitas(_sig=None) -> pd.DataFrame:
    """Carga Operaciones y deja ``lat``/``lon`` listos."""
    _file_signature(VISITAS_PATH)
    df = pd.read_excel(
        VISITAS_PATH, sheet_name="Visitas_Operaciones", engine="openpyxl"
    )
    df.columns = [str(c) for c in df.columns]
    df = df.rename(columns=VISITAS_RENAME)
    df.columns = [c.strip() if isinstance(c, str) else c for c in df.columns]

    if "Nombre del Punto" in df.columns:
        df["Nombre del Punto"] = df["Nombre del Punto"].astype(str).str.strip()
        df = df[df["Nombre del Punto"].ne("") & df["Nombre del Punto"].ne("nan")]

    fecha_col = _find_date_column(df)
    if fecha_col is not None:
        df[DATE_COLUMN_STD] = pd.to_datetime(
            df[fecha_col], errors="coerce", dayfirst=True
        )
    else:
        df[DATE_COLUMN_STD] = pd.NaT

    if "ID" in df.columns:
        df["ID"] = df["ID"].astype(str)
    elif "Nombre del Punto" in df.columns:
        df["ID"] = df["Nombre del Punto"]
    else:
        df["ID"] = df.index.astype(str)

    maps_col, address_col = _add_visit_coordinates(df)
    df.attrs["coordinate_sources"] = {
        "maps_column": maps_col,
        "address_column": address_col,
        "con_xy_y_x": ("Y" in df.columns and "X" in df.columns),
    }
    return df.reset_index(drop=True)


def reload_all():
    """Fuerza la recarga de los Excel desde la interfaz."""
    load_tiendas.clear()
    load_visitas.clear()
